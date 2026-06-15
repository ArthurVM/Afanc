from __future__ import annotations

import json
from collections import Counter
from os import path, walk
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .taxadd import FASTA_EXTS, normaliseTaxonName


def write_autodatabase_manifest(args: Any) -> Path:
    """Write a human-readable XLSX manifest for an Afanc autodatabase run."""
    records = build_manifest_records(args)
    workbook_path = Path(args.autoDB_WDir) / f"{args.output_prefix}.manifest.xlsx"
    write_xlsx(
        workbook_path,
        {
            "Assemblies": records,
            "Taxa summary": summarise_taxa(records),
            "Run summary": summarise_run(args, records),
        },
    )
    return workbook_path


def build_manifest_records(args: Any) -> List[Dict[str, Any]]:
    fasta_dir = Path(args.fastaDir)
    mappings = _load_json(Path(args.autoDB_WDir) / "taxID_mappings.json", default={})
    db_fastas = _load_json(Path(args.autoDB_WDir) / "fastas_in_DB.json", default={})
    included_by_taxid = {
        str(taxid): {str(fasta_name) for fasta_name in fasta_names}
        for taxid, fasta_names in db_fastas.items()
    }
    taxid_by_name = {
        normaliseTaxonName(taxname): str(taxid)
        for taxid, taxname in mappings.items()
    }
    mash_metrics = read_mash_metrics(Path(args.qc_WDir))
    warning_taxids = read_warning_taxids(Path(args.qc_WDir))

    records = []
    for fasta_path in sorted(_input_fastas(fasta_dir), key=lambda item: str(item)):
        relative_dir = fasta_path.parent.relative_to(fasta_dir)
        parent_taxon = ""
        if len(relative_dir.parts) > 1:
            parent_taxon = normaliseTaxonName(relative_dir.parts[-2])
        taxon_name = normaliseTaxonName(fasta_path.parent.name)
        taxid = taxid_by_name.get(taxon_name)
        original_name = fasta_path.name
        rewritten_name = None
        final_name = None
        final_path = None
        in_final_db = False
        reason = "taxon_not_resolved"
        avg_mash_distance = None
        qc_status = "not_run"

        if taxid is not None:
            rewritten_name = _rewritten_fasta_name(original_name, taxid)
            final_name = rewritten_name
            in_final_db = rewritten_name in included_by_taxid.get(taxid, set())
            final_path = str(Path(args.cleanFasta_WDir) / rewritten_name) if in_final_db else ""
            avg_mash_distance = mash_metrics.get(rewritten_name)

            if taxid in warning_taxids:
                qc_status = "not_run_fewer_than_3_assemblies"
                reason = "included_without_mash_qc_fewer_than_3_assemblies" if in_final_db else "not_in_final_db"
            elif in_final_db:
                qc_status = "passed_mash_qc"
                reason = "included_in_final_db"
            else:
                qc_status = "failed_mash_qc"
                reason = "rejected_by_mash_qc"

        records.append(
            {
                "original_fasta": str(fasta_path),
                "original_filename": original_name,
                "input_relative_dir": "" if str(relative_dir) == "." else str(relative_dir),
                "input_taxon": taxon_name,
                "parent_taxon": parent_taxon,
                "taxid": taxid or "",
                "rewritten_fasta": rewritten_name or "",
                "final_db_fasta": final_name or "",
                "in_final_db": "yes" if in_final_db else "no",
                "decision_reason": reason,
                "qc_status": qc_status,
                "average_mash_distance": avg_mash_distance if avg_mash_distance is not None else "",
                "final_db_path": final_path or "",
            }
        )

    return records


def summarise_taxa(records: Iterable[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    summary = {}
    for record in records:
        key = (record["input_taxon"], record["taxid"])
        item = summary.setdefault(
            key,
            {
                "input_taxon": record["input_taxon"],
                "taxid": record["taxid"],
                "provided_assemblies": 0,
                "included_assemblies": 0,
                "rejected_assemblies": 0,
            },
        )
        item["provided_assemblies"] += 1
        if record["in_final_db"] == "yes":
            item["included_assemblies"] += 1
        else:
            item["rejected_assemblies"] += 1

    return sorted(summary.values(), key=lambda item: (str(item["input_taxon"]), str(item["taxid"])))


def summarise_run(args: Any, records: Iterable[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    records = list(records)
    reasons = Counter(record["decision_reason"] for record in records)
    rows = [
        {"field": "output_prefix", "value": args.output_prefix},
        {"field": "input_fasta_dir", "value": str(args.fastaDir)},
        {"field": "autodatabase_dir", "value": str(args.autoDB_WDir)},
        {"field": "ncbi_date", "value": getattr(args, "ncbi_date", "")},
        {"field": "mode_range", "value": getattr(args, "mode_range", "")},
        {"field": "provided_assemblies", "value": len(records)},
        {"field": "included_assemblies", "value": sum(record["in_final_db"] == "yes" for record in records)},
        {"field": "rejected_assemblies", "value": sum(record["in_final_db"] != "yes" for record in records)},
    ]
    rows.extend(
        {"field": f"decision_reason:{reason}", "value": count}
        for reason, count in sorted(reasons.items())
    )
    return rows


def read_mash_metrics(qc_dir: Path) -> Dict[str, float]:
    metrics = {}
    if not qc_dir.is_dir():
        return metrics

    for mash_file in qc_dir.glob("*_mash.txt"):
        with mash_file.open("r") as handle:
            for line in handle:
                parts = line.strip().split()
                if len(parts) < 2:
                    continue
                try:
                    metrics[path.basename(parts[0])] = float(parts[-1])
                except ValueError:
                    continue
    return metrics


def read_warning_taxids(qc_dir: Path) -> set[str]:
    if not qc_dir.is_dir():
        return set()
    return {
        warning_path.name.split("_warning.txt")[0]
        for warning_path in qc_dir.glob("*_warning.txt")
    }


def write_xlsx(workbook_path: Path, sheets: Mapping[str, List[Mapping[str, Any]]]) -> Path:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet_name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=_safe_sheet_name(sheet_name))
        headers = list(rows[0].keys()) if rows else ["message"]
        data_rows = rows if rows else [{"message": "No records"}]
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in data_rows:
            worksheet.append([row.get(header, "") for header in headers])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        _set_column_widths(worksheet, headers, data_rows)

    workbook.save(workbook_path)
    return workbook_path


def _input_fastas(fasta_dir: Path) -> Iterable[Path]:
    for dirpath, _subdirs, filenames in walk(fasta_dir):
        for filename in filenames:
            if filename.endswith(FASTA_EXTS):
                yield Path(dirpath) / filename


def _rewritten_fasta_name(original_name: str, taxid: str) -> str:
    if original_name.endswith(".gz"):
        original_name = original_name.rsplit(".gz", 1)[0]
    return f"{taxid}_{original_name}"


def _load_json(json_path: Path, default: Any) -> Any:
    if not json_path.is_file():
        return default
    with json_path.open("r") as handle:
        return json.load(handle)


def _set_column_widths(worksheet: Any, headers: List[str], rows: List[Mapping[str, Any]]) -> None:
    for column_index, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            max_len = max(max_len, len(str(row.get(header, ""))))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_len + 2, 12), 60)


def _safe_sheet_name(name: str) -> str:
    safe = "".join("_" if char in "[]:*?/\\'" else char for char in str(name))
    return safe[:31] or "Sheet"
