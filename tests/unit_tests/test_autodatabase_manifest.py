import json
from types import SimpleNamespace

from openpyxl import load_workbook

from Afanc.autodatabase.manifest import build_manifest_records, write_autodatabase_manifest


def _make_args(tmp_path):
    input_dir = tmp_path / "input_fastas"
    auto_db = tmp_path / "Afanc_autodb"
    qc_dir = auto_db / "selectFasta_autoDatabase_qc"
    clean_dir = auto_db / "selectFasta_autoDatabase_cleanFasta"
    input_dir.mkdir()
    auto_db.mkdir()
    qc_dir.mkdir()
    clean_dir.mkdir()
    return SimpleNamespace(
        fastaDir=input_dir,
        autoDB_WDir=auto_db,
        qc_WDir=qc_dir,
        cleanFasta_WDir=clean_dir,
        output_prefix="Afanc_autodb",
        ncbi_date="2026-05-01",
        mode_range=0.1,
    )


def test_build_manifest_records_reports_included_rejected_and_unresolved(tmp_path):
    args = _make_args(tmp_path)
    species_dir = args.fastaDir / "Mycobacterium_tuberculosis"
    unresolved_dir = args.fastaDir / "Unknown_species"
    species_dir.mkdir()
    unresolved_dir.mkdir()
    (species_dir / "good.fna").write_text(">a\nACGT\n")
    (species_dir / "bad.fna").write_text(">a\nACGT\n")
    (unresolved_dir / "unknown.fna").write_text(">a\nACGT\n")
    (args.cleanFasta_WDir / "1773_good.fna").write_text(">a\nACGT\n")
    (args.autoDB_WDir / "taxID_mappings.json").write_text(
        json.dumps({"1773": "Mycobacterium tuberculosis"})
    )
    (args.autoDB_WDir / "fastas_in_DB.json").write_text(
        json.dumps({"1773": ["1773_good.fna"]})
    )
    (args.qc_WDir / "1773_mash.txt").write_text(
        f"{args.autoDB_WDir}/selectFasta_autoDatabase_Fasta/1773_good.fna 0 0.01\n"
        f"{args.autoDB_WDir}/selectFasta_autoDatabase_Fasta/1773_bad.fna 0 0.42\n"
    )

    records = build_manifest_records(args)
    by_name = {record["original_filename"]: record for record in records}

    assert by_name["good.fna"]["in_final_db"] == "yes"
    assert by_name["good.fna"]["input_relative_dir"] == "Mycobacterium_tuberculosis"
    assert by_name["good.fna"]["decision_reason"] == "included_in_final_db"
    assert by_name["good.fna"]["average_mash_distance"] == 0.01
    assert by_name["bad.fna"]["in_final_db"] == "no"
    assert by_name["bad.fna"]["decision_reason"] == "rejected_by_mash_qc"
    assert by_name["unknown.fna"]["taxid"] == ""
    assert by_name["unknown.fna"]["decision_reason"] == "taxon_not_resolved"


def test_manifest_marks_small_taxa_as_included_without_qc(tmp_path):
    args = _make_args(tmp_path)
    species_dir = args.fastaDir / "Mycobacterium_avium"
    species_dir.mkdir()
    (species_dir / "only.fna.gz").write_text("not really gzipped but name is enough")
    (args.cleanFasta_WDir / "1764_only.fna").write_text(">a\nACGT\n")
    (args.autoDB_WDir / "taxID_mappings.json").write_text(json.dumps({"1764": "Mycobacterium avium"}))
    (args.autoDB_WDir / "fastas_in_DB.json").write_text(json.dumps({"1764": ["1764_only.fna"]}))
    (args.qc_WDir / "1764_warning.txt").write_text("Warning: Taxon 1764 has fewer than 3 samples.")

    record = build_manifest_records(args)[0]

    assert record["rewritten_fasta"] == "1764_only.fna"
    assert record["in_final_db"] == "yes"
    assert record["qc_status"] == "not_run_fewer_than_3_assemblies"
    assert record["decision_reason"] == "included_without_mash_qc_fewer_than_3_assemblies"


def test_write_autodatabase_manifest_creates_xlsx_workbook(tmp_path):
    args = _make_args(tmp_path)
    species_dir = args.fastaDir / "Mycobacterium_tuberculosis"
    species_dir.mkdir()
    (species_dir / "good.fna").write_text(">a\nACGT\n")
    (args.cleanFasta_WDir / "1773_good.fna").write_text(">a\nACGT\n")
    (args.autoDB_WDir / "taxID_mappings.json").write_text(json.dumps({"1773": "Mycobacterium tuberculosis"}))
    (args.autoDB_WDir / "fastas_in_DB.json").write_text(json.dumps({"1773": ["1773_good.fna"]}))

    workbook_path = write_autodatabase_manifest(args)

    assert workbook_path == args.autoDB_WDir / "Afanc_autodb.manifest.xlsx"
    assert workbook_path.exists()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert workbook.sheetnames == ["Assemblies", "Taxa summary", "Run summary"]
    worksheet = workbook["Assemblies"]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    assert "original_fasta" in headers
    assert "decision_reason" in headers
    assert any("included_in_final_db" in row for row in rows)
